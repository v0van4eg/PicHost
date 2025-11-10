# app.py

from auth_system import AuthManager, login_required, admin_required, role_required, auth_context_processor, \
    is_authenticated, get_current_user
import os
import zipfile
from flask import Flask, request, session, jsonify, render_template, send_from_directory, send_file
import logging
from urllib.parse import quote
from PIL import Image
import io
import hashlib
import shutil
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import tempfile
import atexit
import json
from database import db_manager
from werkzeug.middleware.proxy_fix import ProxyFix
from zip_processor import ZipProcessor
from utils import safe_folder_name, cleanup_album_thumbnails, cleanup_empty_folders
from sync_manager import SyncManager

app = Flask(__name__)
app.secret_key = os.environ.get('FLASK_SECRET_KEY', 'default_secret_key')
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_prefix=1)

# Инициализация аутентификации (теперь параметры берутся из переменных окружения)
auth_manager = AuthManager()
auth_manager.init_app(app)

# Регистрация маршрутов аутентификации
auth_manager.register_routes()

# Добавление контекстного процессора
app.context_processor(auth_context_processor)

app.config['UPLOAD_FOLDER'] = 'images'
app.config['THUMBNAIL_FOLDER'] = 'thumbnails'
app.config['THUMBNAIL_SIZE'] = (120, 120)  # Размер превью
app.config['PREVIEW_SIZE'] = (400, 400)  # Размер для предпросмотра
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024 * 1024  # 16GB

# Создаем папки если их нет
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['THUMBNAIL_FOLDER'], exist_ok=True)

# Логирование
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

# Конфигурация домена и базового URL получаем из переменных окружения
domain = os.environ.get('DOMAIN', 'pichosting.mooo.com')
base_url = f"http://{domain}"

# Инициализация ZipProcessor
zip_processor = ZipProcessor(
    upload_folder=app.config['UPLOAD_FOLDER'],
    base_url=base_url,
    thumbnail_folder=app.config['THUMBNAIL_FOLDER']
)

sync_manager = SyncManager(
    upload_folder=app.config['UPLOAD_FOLDER'],
    base_url=base_url,
    thumbnail_folder=app.config['THUMBNAIL_FOLDER']
)


def generate_image_hash(file_path):
    """Генерирует хэш для файла для кэширования"""
    try:
        with open(file_path, 'rb') as f:
            return hashlib.md5(f.read()).hexdigest()
    except Exception as e:
        logger.error(f"Error generating hash for {file_path}: {e}")
        return hashlib.md5(file_path.encode()).hexdigest()


def create_thumbnail(original_path, size, quality=85):
    """Создает миниатюру изображения"""
    try:
        with Image.open(original_path) as img:
            if img.mode in ('RGBA', 'P'):
                img = img.convert('RGB')

            img.thumbnail(size, Image.Resampling.LANCZOS)

            buffer = io.BytesIO()
            img.save(buffer, 'JPEG', quality=quality, optimize=True)
            buffer.seek(0)

            return buffer
    except Exception as e:
        logger.error(f"Error creating thumbnail for {original_path}: {e}")
        return None


def get_thumbnail_path(original_path, size):
    """Генерирует путь для миниатюры"""
    file_hash = generate_image_hash(original_path)
    filename = os.path.basename(original_path)
    name, ext = os.path.splitext(filename)
    size_str = f"{size[0]}x{size[1]}"

    rel_path = os.path.relpath(original_path, app.config['UPLOAD_FOLDER'])
    rel_dir = os.path.dirname(rel_path)

    thumbnail_filename = f"{name}_{size_str}_{file_hash[:8]}.jpg"

    if rel_dir and rel_dir != '.':
        thumbnail_dir = os.path.join(app.config['THUMBNAIL_FOLDER'], rel_dir)
        os.makedirs(thumbnail_dir, exist_ok=True)
        return os.path.join(thumbnail_dir, thumbnail_filename)
    else:
        return os.path.join(app.config['THUMBNAIL_FOLDER'], thumbnail_filename)


def cleanup_file_thumbnails(filename):
    """Очищает превью для конкретного файла"""
    try:
        original_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        if not os.path.exists(original_path):
            # Если оригинального файла нет, ищем и удаляем все возможные превью
            rel_dir = os.path.dirname(filename)
            file_base = os.path.splitext(os.path.basename(filename))[0]

            if rel_dir and rel_dir != '.':
                thumb_dir = os.path.join(app.config['THUMBNAIL_FOLDER'], rel_dir)
                if os.path.exists(thumb_dir):
                    # Удаляем все превью для этого файла
                    for thumb_file in os.listdir(thumb_dir):
                        if thumb_file.startswith(file_base + '_'):
                            thumb_path = os.path.join(thumb_dir, thumb_file)
                            os.remove(thumb_path)
                            logger.info(f"Deleted orphaned thumbnail: {thumb_path}")
        else:
            # Удаляем превью для существующего файла
            thumbnail_path = get_thumbnail_path(original_path, app.config['THUMBNAIL_SIZE'])
            preview_path = get_thumbnail_path(original_path, app.config['PREVIEW_SIZE'])

            for thumb_path in [thumbnail_path, preview_path]:
                if os.path.exists(thumb_path):
                    os.remove(thumb_path)
                    logger.info(f"Deleted thumbnail: {thumb_path}")

    except Exception as e:
        logger.error(f"Error cleaning up thumbnails for file {filename}: {e}")


# Инициализация базы данных
def init_db():
    """Инициализация базы данных при запуске приложения"""
    max_retries = 5
    retry_delay = 2

    for attempt in range(max_retries):
        try:
            # Проверяем соединение и существование таблицы
            result = db_manager.execute_query("""
                SELECT EXISTS (
                    SELECT FROM information_schema.tables 
                    WHERE table_name = 'files'
                );
            """, fetch=True)

            table_exists = result[0]['exists'] if result else False

            if not table_exists:
                logger.warning("Table 'files' does not exist. Creating...")
                db_manager.execute_query('''
                    CREATE TABLE files (
                        id SERIAL PRIMARY KEY,
                        filename TEXT NOT NULL,
                        album_name TEXT NOT NULL,
                        article_number TEXT NOT NULL,
                        public_link TEXT NOT NULL,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                    )
                ''', commit=True)

                # Создаем индексы
                db_manager.execute_query('CREATE INDEX idx_files_album_name ON files(album_name)', commit=True)
                db_manager.execute_query('CREATE INDEX idx_files_article_number ON files(article_number)', commit=True)
                db_manager.execute_query('CREATE INDEX idx_files_created_at ON files(created_at)', commit=True)

                logger.info("Table 'files' created successfully")

            # Проверяем существование таблицы логов действий пользователей
            result_logs = db_manager.execute_query("""
                SELECT EXISTS (
                    SELECT FROM information_schema.tables 
                    WHERE table_name = 'user_actions_log'
                );
            """, fetch=True)

            logs_table_exists = result_logs[0]['exists'] if result_logs else False

            if not logs_table_exists:
                logger.warning("Table 'user_actions_log' does not exist. Creating...")
                db_manager.execute_query('''
                    CREATE TABLE user_actions_log (
                        id SERIAL PRIMARY KEY,
                        user_id TEXT NOT NULL,
                        username TEXT NOT NULL,
                        action TEXT NOT NULL,
                        resource_type TEXT,
                        resource_name TEXT,
                        details JSONB,
                        timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                    )
                ''', commit=True)

                # Создаем индексы для логов
                db_manager.execute_query('CREATE INDEX idx_user_actions_user_id ON user_actions_log(user_id)',
                                         commit=True)
                db_manager.execute_query('CREATE INDEX idx_user_actions_action ON user_actions_log(action)',
                                         commit=True)
                db_manager.execute_query('CREATE INDEX idx_user_actions_timestamp ON user_actions_log(timestamp)',
                                         commit=True)

                logger.info("Table 'user_actions_log' created successfully")

            logger.info("Database initialized successfully")
            return

        except Exception as e:
            logger.warning(f"Database initialization attempt {attempt + 1} failed: {e}")
            if attempt < max_retries - 1:
                import time
                time.sleep(retry_delay)
            else:
                logger.error(f"Failed to initialize database after {max_retries} attempts: {e}")
                raise


# Получение списка альбомов
def get_albums():
    results = db_manager.execute_query("SELECT DISTINCT album_name FROM files", fetch=True)
    return [album['album_name'] for album in results] if results else []


# Получение списка артикулов для указанного альбома
def get_articles(album_name):
    results = db_manager.execute_query(
        "SELECT DISTINCT article_number FROM files WHERE album_name = %s",
        (album_name,),
        fetch=True
    )
    return [article['article_number'] for article in results] if results else []


# Получение всех файлов из БД
def get_all_files():
    results = db_manager.execute_query(
        "SELECT filename, album_name, article_number, public_link, created_at FROM files ORDER BY created_at DESC",
        fetch=True
    )
    return results if results else []


# Синхронизация БД с файловой системой
# def sync_db_with_filesystem():
#     """
#     Синхронизирует базу данных с файловой системой в одной транзакции.
#     """
#     try:
#         # Получаем все файлы из БД в одном запросе
#         db_files_result = db_manager.execute_query(
#             "SELECT filename, album_name, article_number, public_link FROM files",
#             fetch=True
#         )
#         db_files = {row['filename']: {
#             'album_name': row['album_name'],
#             'article_number': row['article_number'],
#             'public_link': row['public_link']
#         } for row in db_files_result} if db_files_result else {}
#
#         # Сканируем файловую систему
#         fs_files = {}
#         allowed_extensions = {'.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp', '.tiff', '.svg'}
#
#         for root, dirs, files in os.walk(app.config['UPLOAD_FOLDER']):
#             for file in files:
#                 _, ext = os.path.splitext(file.lower())
#                 if ext in allowed_extensions:
#                     full_path = os.path.join(root, file)
#                     rel_path = os.path.relpath(full_path, app.config['UPLOAD_FOLDER']).replace(os.sep, '/')
#
#                     # Определяем альбом и артикул из пути
#                     path_parts = rel_path.split('/')
#                     if len(path_parts) >= 1:
#                         album_name = path_parts[0]
#
#                         # Если файл находится в подпапке (артикуле)
#                         if len(path_parts) >= 3:
#                             article_number = path_parts[1]
#                         else:
#                             # Если файл напрямую в альбоме, используем имя файла без расширения как артикул
#                             article_number = os.path.splitext(file)[0]
#
#                         # Обеспечиваем безопасные имена
#                         album_name = safe_folder_name(album_name)
#                         article_number = safe_folder_name(article_number)
#
#                         encoded_path = quote(rel_path, safe='/')
#                         public_link = f"{base_url}/images/{encoded_path}"
#
#                         fs_files[rel_path] = {
#                             'album_name': album_name,
#                             'article_number': article_number,
#                             'public_link': public_link
#                         }
#
#         # Находим файлы для удаления и добавления
#         files_to_delete = set(db_files.keys()) - set(fs_files.keys())
#         files_to_add = set(fs_files.keys()) - set(db_files.keys())
#
#         # Подготавливаем операции для транзакции
#         operations = []
#
#         # Операция удаления
#         if files_to_delete:
#             delete_query = "DELETE FROM files WHERE filename = ANY(%s)"
#             operations.append((delete_query, (list(files_to_delete),)))
#
#         # Операция вставки
#         if files_to_add:
#             insert_data = []
#             for rel_path in files_to_add:
#                 file_info = fs_files[rel_path]
#                 insert_data.append((
#                     rel_path,
#                     file_info['album_name'],
#                     file_info['article_number'],
#                     file_info['public_link']
#                 ))
#
#             insert_query = """
#                 INSERT INTO files (filename, album_name, article_number, public_link)
#                 VALUES (%s, %s, %s, %s)
#             """
#             operations.append((insert_query, insert_data, True))  # True для executemany
#
#         # Выполняем все операции в одной транзакции
#         if operations:
#             db_manager.execute_in_transaction(operations)
#
#         # Очищаем превью для удаленных файлов (вне транзакции, так как это файловые операции)
#         for rel_path in files_to_delete:
#             cleanup_file_thumbnails(rel_path)
#
#         logger.info(f"Sync completed: deleted {len(files_to_delete)} records, added {len(files_to_add)} records")
#         return list(files_to_delete), list(files_to_add)
#
#     except Exception as e:
#         logger.error(f"Error in sync_db_with_filesystem: {e}")
#         raise


# Эндпоинт синхронизации БД (обновленный)
@app.route('/api/sync', methods=['GET'])
@login_required
def api_sync():
    try:
        deleted, added = sync_manager.sync()
        return jsonify({
            'message': 'Synchronization completed successfully',
            'deleted': deleted,
            'added': added
        })
    except Exception as e:
        logger.error(f"Error in sync endpoint: {e}")
        return jsonify({'error': f'Synchronization failed: {str(e)}'}), 500


# Новый эндпоинт для статистики синхронизации
@app.route('/api/sync/stats', methods=['GET'])
@login_required
def api_sync_stats():
    """Возвращает статистику синхронизации"""
    try:
        stats = sync_manager.get_sync_stats()
        return jsonify(stats)
    except Exception as e:
        logger.error(f"Error getting sync stats: {e}")
        return jsonify({'error': str(e)}), 500


def log_user_action(action, resource_type=None, resource_name=None, details=None):
    """
    Записывает действие пользователя в базу данных.

    :param action: str - Тип действия (например, 'upload', 'delete_album', 'delete_article')
    :param resource_type: str - Тип ресурса ('file', 'album', 'article')
    :param resource_name: str - Имя ресурса
    :param details: dict - Дополнительные детали (будет сериализовано в JSON)
    """
    user = get_current_user()
    if not user:
        # Если пользователь не аутентифицирован, можно логировать как анонимное действие
        # или использовать специальное имя/ID.
        # В данном примере используем 'anonymous'
        user_id = 'anonymous'
        username = 'anonymous'
    else:
        user_id = user.get('sub')  # Используем уникальный идентификатор пользователя из OIDC
        username = user.get('name', user.get('preferred_username', 'unknown_user'))

    details_json = json.dumps(details) if details else None

    query = """
    INSERT INTO user_actions_log (user_id, username, action, resource_type, resource_name, details)
    VALUES (%s, %s, %s, %s, %s, %s)
    """
    try:
        db_manager.execute_query(query, (user_id, username, action, resource_type, resource_name, details_json),
                                 commit=True)
        logger.info(
            f"Logged action '{action}' for user '{username}' on {resource_type or 'N/A'} '{resource_name or 'N/A'}'")
    except Exception as e:
        logger.error(f"Failed to log action '{action}' for user '{username}': {e}")


# --- Routes ---
@app.route('/')
def index():
    # Используем функцию из контекстного процессора
    if is_authenticated():
        return render_template('index.html', base_url=base_url)
    else:
        return render_template('hello.html')


@app.route('/hello')
def hello():
    return render_template('hello.html', base_url=base_url)


# Эндпоинт для принудительной очистки превью альбома
@app.route('/api/cleanup-thumbnails/<album_name>', methods=['POST'])
@login_required
def api_cleanup_thumbnails(album_name):
    """Принудительная очистка превью для альбома"""
    try:
        cleanup_album_thumbnails(album_name, app.config['THUMBNAIL_FOLDER'])
        return jsonify({'message': f'Thumbnails for album {album_name} cleaned up successfully'})
    except Exception as e:
        logger.error(f"Error cleaning up thumbnails for {album_name}: {e}")
        return jsonify({'error': f'Cleanup failed: {str(e)}'}), 500


# Загрузка ZIP
@app.route('/upload', methods=['POST'])
@login_required
def upload_zip():
    logger.info("Upload endpoint called")
    if 'zipfile' not in request.files:
        return jsonify({'error': 'No file part'}), 400

    file = request.files['zipfile']

    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400

    if file:
        original_name = file.filename
        base_name = os.path.basename(original_name)
        name_without_ext, _ = os.path.splitext(base_name)
        safe_zip_name = safe_folder_name(name_without_ext) + '.zip'
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], safe_zip_name)
        file.save(file_path)

        # Используем zip_processor для обработки ZIP файла
        success, result = zip_processor.process_zip(file_path)

        if success:
            os.remove(file_path)
            log_user_action('upload', 'album', safe_zip_name)
            return jsonify({'message': 'Files uploaded successfully', 'album_name': result})
        else:
            # Удаляем ZIP файл в случае ошибки
            os.remove(file_path)
            return jsonify({'error': f'Failed to process ZIP file: {result}'}), 500


# API: список всех файлов
@app.route('/api/files')
@login_required
def api_files():
    logger.info("API files endpoint called")
    files = get_all_files()
    return jsonify(files)


# API: список альбомов
@app.route('/api/albums')
@login_required
def api_albums():
    logger.info("API albums endpoint called")
    albums = get_albums()
    return jsonify(albums)


# API: список артикулов для альбома
@app.route('/api/articles/<album_name>')
@login_required
def api_articles(album_name):
    logger.info(f"API articles endpoint called for album: {album_name}")
    articles = get_articles(album_name)
    return jsonify(articles)


# API: получение файлов для конкретного альбома (и опционально артикула)
@app.route('/api/files/<album_name>')
@app.route('/api/files/<album_name>/<article_name>')
@login_required
def api_files_filtered(album_name, article_name=None):
    logger.info(f"API files filtered endpoint called for album: {album_name}, article: {article_name}")
    if article_name:
        results = db_manager.execute_query(
            "SELECT filename, album_name, article_number, public_link, created_at FROM files WHERE album_name = %s AND article_number = %s ORDER BY created_at DESC",
            (album_name, article_name),
            fetch=True
        )
    else:
        results = db_manager.execute_query(
            "SELECT filename, album_name, article_number, public_link, created_at FROM files WHERE album_name = %s ORDER BY created_at DESC",
            (album_name,),
            fetch=True
        )

    return jsonify(results if results else [])


# Новые эндпоинты для превью
@app.route('/api/thumbnails/<album_name>')
@app.route('/api/thumbnails/<album_name>/<article_name>')
@login_required
def api_thumbnails(album_name, article_name=None):
    """API для получения информации о файлах с превью"""
    try:
        if article_name:
            results = db_manager.execute_query(
                """SELECT filename, album_name, article_number, public_link, created_at 
                   FROM files WHERE album_name = %s AND article_number = %s 
                   ORDER BY created_at DESC""",
                (album_name, article_name),
                fetch=True
            )
        else:
            results = db_manager.execute_query(
                """SELECT filename, album_name, article_number, public_link, created_at 
                   FROM files WHERE album_name = %s 
                   ORDER BY created_at DESC""",
                (album_name,),
                fetch=True
            )

        files_data = []
        if results:
            for row in results:
                filename = row['filename']
                album = row['album_name']
                article = row['article_number']
                public_link = row['public_link']
                created_at = row['created_at']

                original_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)

                files_data.append({
                    'filename': filename,
                    'album_name': album,
                    'article_number': article,
                    'public_link': public_link,
                    'created_at': created_at,
                    'thumbnail_url': f"/thumbnails/small/{filename}",
                    'preview_url': f"/thumbnails/medium/{filename}",
                    'file_size': os.path.getsize(original_path) if os.path.exists(original_path) else 0
                })

        return jsonify(files_data)

    except Exception as e:
        logger.error(f"Error in api_thumbnails: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/thumbnails/small/<path:filename>')
@login_required
def serve_small_thumbnail(filename):
    """Отдает маленькие превью (120x120)"""
    return serve_thumbnail(filename, app.config['THUMBNAIL_SIZE'])


@app.route('/thumbnails/medium/<path:filename>')
@login_required
def serve_medium_thumbnail(filename):
    """Отдает средние превью (400x400)"""
    return serve_thumbnail(filename, app.config['PREVIEW_SIZE'])


def serve_thumbnail(filename, size):
    """Обслуживает миниатюры, создавая их при необходимости"""
    original_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)

    if not os.path.exists(original_path):
        return jsonify({'error': 'File not found'}), 404

    thumbnail_path = get_thumbnail_path(original_path, size)

    # Создаем миниатюру если ее нет
    if not os.path.exists(thumbnail_path):
        thumbnail_buffer = create_thumbnail(original_path, size)
        if thumbnail_buffer:
            with open(thumbnail_path, 'wb') as f:
                f.write(thumbnail_buffer.getvalue())
            logger.info(f"Created new thumbnail: {thumbnail_path}")
        else:
            return send_from_directory('static', 'image-placeholder.png')

    return send_from_directory(os.path.dirname(thumbnail_path),
                               os.path.basename(thumbnail_path))


@app.route('/api/export-xlsx', methods=['POST'])
@login_required
def api_export_xlsx():
    """Создание XLSX документа с ссылками"""
    logger.info("API export XLSX endpoint called")
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400

        album_name = data.get('album_name')
        article_name = data.get('article_name')  # Может быть None для всех артикулов
        export_type = data.get('export_type')  # 'in_row' или 'in_cell'
        separator = data.get('separator', ', ')  # Разделитель для варианта "в ячейку"

        if not album_name or not export_type:
            return jsonify({'error': 'Missing required parameters'}), 400

        # Получаем данные из БД
        if article_name:
            results = db_manager.execute_query(
                "SELECT filename, article_number, public_link FROM files WHERE album_name = %s AND article_number = %s ORDER BY article_number, filename",
                (album_name, article_name),
                fetch=True
            )
        else:
            results = db_manager.execute_query(
                "SELECT filename, article_number, public_link FROM files WHERE album_name = %s ORDER BY article_number, filename",
                (album_name,),
                fetch=True
            )

        if not results:
            return jsonify({'error': 'No data found for export'}), 404

        # Функция для извлечения числового суффикса из имени файла
        def extract_suffix(filename):
            import re
            # Ищем паттерн: любое количество символов, затем подчеркивание, затем цифры до точки
            match = re.search(r'(.+)_(\d+)(\.[^.]*)?$', filename)
            if match:
                return int(match.group(2))  # Возвращаем числовое значение
            return 0  # Если суффикс не найден

        # Сортируем результаты по артикулу и числовому суффиксу в имени файла
        sorted_results = sorted(results, key=lambda x: (x['article_number'], extract_suffix(x['filename'])))

        # Группируем ссылки по артикулам с правильной сортировкой
        articles_data = {}
        for row in sorted_results:
            article = row['article_number']
            if article not in articles_data:
                articles_data[article] = []
            articles_data[article].append(row['public_link'])

        # Создаем Excel файл
        wb = Workbook()
        ws = wb.active
        ws.title = "Ссылки на изображения"

        # Стили для шапки
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        if export_type == 'in_row':
            # Вариант "В строку"
            # Определяем максимальное количество ссылок
            max_links = max(len(links) for links in articles_data.values())

            # Создаем шапку
            headers = ['Артикул'] + [f'Ссылка {i + 1}' for i in range(max_links)]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill

            # Заполняем данные (уже отсортированные)
            for row, (article, links) in enumerate(articles_data.items(), 2):
                ws.cell(row=row, column=1, value=article)
                for col, link in enumerate(links, 2):
                    ws.cell(row=row, column=col, value=link)

        elif export_type == 'in_cell':
            # Вариант "В ячейку"
            headers = ['Артикул', 'Ссылки']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill

            # Заполняем данные (уже отсортированные)
            for row, (article, links) in enumerate(articles_data.items(), 2):
                ws.cell(row=row, column=1, value=article)
                # Объединяем ссылки через разделитель (уже отсортированные)
                links_text = separator.join(links)
                ws.cell(row=row, column=2, value=links_text)

        # Авто-ширина колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Сохраняем файл во временную директорию
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            wb.save(tmp_file.name)
            tmp_filename = tmp_file.name

        # Генерируем имя файла
        filename = f"links_{album_name}"
        if article_name:
            filename += f"_{article_name}"
        filename += ".xlsx"

        # Отправляем файл
        response = send_file(
            tmp_filename,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        # Удаляем временный файл после отправки
        @response.call_on_close
        def remove_temp_file():
            try:
                os.unlink(tmp_filename)
            except Exception as e:
                logger.error(f"Error removing temporary file {tmp_filename}: {e}")

        return response

    except Exception as e:
        logger.error(f"Error creating XLSX file: {e}")
        return jsonify({'error': f'Failed to create XLSX file: {str(e)}'}), 500


@app.route('/api/delete-album/<album_name>', methods=['DELETE'])
@login_required
@role_required(['appadmin'])
def api_delete_album(album_name):
    """Удаление альбома из БД и файловой системы"""
    logger.info(f"API delete album endpoint called for: {album_name}")
    try:
        # Получаем все файлы альбома
        files = db_manager.execute_query(
            "SELECT filename FROM files WHERE album_name = %s",
            (album_name,),
            fetch=True
        )
        filenames = [file['filename'] for file in files] if files else []

        # Удаляем записи из БД
        db_manager.execute_query(
            "DELETE FROM files WHERE album_name = %s",
            (album_name,),
            commit=True
        )

        # Удаляем файлы и папки
        album_path = os.path.join(app.config['UPLOAD_FOLDER'], album_name)
        thumbnail_album_path = os.path.join(app.config['THUMBNAIL_FOLDER'], album_name)

        # Удаляем файлы изображений
        if os.path.exists(album_path):
            shutil.rmtree(album_path)
            logger.info(f"Deleted album directory: {album_path}")

        # Удаляем превью альбома
        cleanup_album_thumbnails(album_name, app.config['THUMBNAIL_FOLDER'])

        # Удаляем папку превью если осталась
        if os.path.exists(thumbnail_album_path):
            shutil.rmtree(thumbnail_album_path)
            logger.info(f"Deleted album thumbnails directory: {thumbnail_album_path}")

        log_user_action('delete_album', 'album', album_name,
                        {'deleted_files_count': len(filenames) if 'filenames' in locals() else 'unknown'})
        return jsonify({'message': f'Альбом "{album_name}" успешно удален'})

    except Exception as e:
        logger.error(f"Error deleting album {album_name}: {e}")
        return jsonify({'error': f'Ошибка удаления альбома: {str(e)}'}), 500


@app.route('/api/delete-article/<album_name>/<article_name>', methods=['DELETE'])
@login_required
@role_required(['appadmin'])
def api_delete_article(album_name, article_name):
    """Удаление артикула из БД и файловой системы"""
    logger.info(f"API delete article endpoint called for: {album_name}/{article_name}")
    try:
        # Получаем все файлы артикула
        files = db_manager.execute_query(
            "SELECT filename FROM files WHERE album_name = %s AND article_number = %s",
            (album_name, article_name),
            fetch=True
        )
        filenames = [file['filename'] for file in files] if files else []

        # Удаляем записи из БД
        db_manager.execute_query(
            "DELETE FROM files WHERE album_name = %s AND article_number = %s",
            (album_name, article_name),
            commit=True
        )

        # Удаляем файлы и папки
        article_path = os.path.join(app.config['UPLOAD_FOLDER'], album_name, article_name)

        # Удаляем файлы изображений
        if os.path.exists(article_path):
            shutil.rmtree(article_path)
            logger.info(f"Deleted article directory: {article_path}")

        # Удаляем превью для каждого файла
        for filename in filenames:
            cleanup_file_thumbnails(filename)

        # Удаляем папку превью артикула если осталась
        thumbnail_article_path = os.path.join(app.config['THUMBNAIL_FOLDER'], album_name, article_name)
        if os.path.exists(thumbnail_article_path):
            shutil.rmtree(thumbnail_article_path)
            logger.info(f"Deleted article thumbnails directory: {thumbnail_article_path}")

        # Синхронизируем БД после удаления
        # sync_manager.sync()
        log_user_action('delete_article', 'article', f"{album_name}/{article_name}",
                        {'deleted_files_count': len(filenames) if 'filenames' in locals() else 'unknown'})
        return jsonify({'message': f'Артикул "{article_name}" в альбоме "{album_name}" успешно удален'})

    except Exception as e:
        logger.error(f"Error deleting article {article_name} from album {album_name}: {e}")
        return jsonify({'error': f'Ошибка удаления артикула: {str(e)}'}), 500


# API: количество файлов в альбоме
@app.route('/api/count/album/<album_name>')
@login_required
def api_count_album(album_name):
    """Возвращает количество файлов в альбоме"""
    logger.info(f"API count album endpoint called for: {album_name}")
    try:
        result = db_manager.execute_query(
            "SELECT COUNT(*) as count FROM files WHERE album_name = %s",
            (album_name,),
            fetch=True
        )
        count = result[0]['count'] if result else 0
        logger.info(f"Album {album_name} has {count} files")
        return jsonify({'count': count})
    except Exception as e:
        logger.error(f"Error counting files for album {album_name}: {e}")
        return jsonify({'error': str(e)}), 500


# API: количество файлов в артикуле
@app.route('/api/count/article/<album_name>/<article_name>')
@login_required
def api_count_article(album_name, article_name):
    """Возвращает количество файлов в артикуле"""
    logger.info(f"API count article endpoint called for: {album_name}/{article_name}")
    try:
        result = db_manager.execute_query(
            "SELECT COUNT(*) as count FROM files WHERE album_name = %s AND article_number = %s",
            (album_name, article_name),
            fetch=True
        )
        count = result[0]['count'] if result else 0
        logger.info(f"Article {article_name} in album {album_name} has {count} files")
        return jsonify({'count': count})
    except Exception as e:
        logger.error(f"Error counting files for article {article_name} in album {album_name}: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/admin')
@login_required
@role_required(['appadmin'])
def admin_panel():
    logger.info("Admin panel accessed")
    user = session.get('user', {})
    display_roles = user.get('display_roles', [])
    all_roles = user.get('roles', [])
    return render_template('admin.html', display_roles=display_roles, all_roles=all_roles)


@app.route('/admin/logs')
@login_required
@admin_required
def admin_logs():
    """Страница с логами действий пользователей"""
    logger.info("Admin logs accessed")
    page = request.args.get('page', 1, type=int)
    per_page = 50  # Количество записей на странице
    offset = (page - 1) * per_page

    search_user = request.args.get('search_user', '')
    search_action = request.args.get('search_action', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')

    # Формируем условия для фильтрации
    conditions = []
    params = []

    if search_user:
        conditions.append("username ILIKE %s")  # ILIKE для нечувствительного к регистру поиска
        params.append(f"%{search_user}%")
    if search_action:
        conditions.append("action = %s")
        params.append(search_action)
    if date_from:
        conditions.append("timestamp >= %s")
        params.append(date_from)
    if date_to:
        conditions.append("timestamp <= %s")
        params.append(date_to)

    where_clause = " WHERE " + " AND ".join(conditions) if conditions else ""
    order_clause = " ORDER BY timestamp DESC LIMIT %s OFFSET %s"
    params.extend([per_page, offset])

    query = f"""
    SELECT user_id, username, action, resource_type, resource_name, details, timestamp
    FROM user_actions_log
    {where_clause}
    {order_clause}
    """

    try:
        logs = db_manager.execute_query(query, tuple(params), fetch=True)
        # Получаем общее количество записей для пагинации
        count_query = f"SELECT COUNT(*) as total FROM user_actions_log {where_clause}"
        total_count_result = db_manager.execute_query(count_query, tuple(params[:-2]),
                                                      fetch=True)  # Исключаем LIMIT и OFFSET
        total_count = total_count_result[0]['total'] if total_count_result else 0
    except Exception as e:
        logger.error(f"Error fetching logs: {e}")
        logs = []
        total_count = 0

    # Получаем уникальные действия для фильтра
    try:
        actions_query = "SELECT DISTINCT action FROM user_actions_log ORDER BY action"
        actions = db_manager.execute_query(actions_query, fetch=True)
        action_list = [row['action'] for row in actions]
    except Exception as e:
        logger.error(f"Error fetching actions for filter: {e}")
        action_list = []

    return render_template('logs.html',
                           logs=logs,
                           page=page,
                           per_page=per_page,
                           total_count=total_count,
                           search_user=search_user,
                           search_action=search_action,
                           date_from=date_from,
                           date_to=date_to,
                           available_actions=action_list,
                           current_user=get_current_user())  # Передаем пользователя в шаблон


# Инициализация базы данных при запуске приложения
init_db()


# Функция для закрытия соединений при выходе
@atexit.register
def cleanup():
    db_manager.close()


# --- Main ---
if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True, ssl_context='adhoc')
