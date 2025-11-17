# document_generator.py

import tempfile
import logging
import csv
# --- ИСПРАВЛЕНО: Добавлен импорт Workbook ---
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from database import db_manager
from utils import log_user_action
from datetime import datetime

logger = logging.getLogger(__name__)


class DocumentGenerator:
    """Класс для генерации различных типов документов (XLSX, CSV, etc.)"""

    def __init__(self, base_url, upload_folder):
        self.base_url = base_url
        self.upload_folder = upload_folder
        # --- Используем константы для стилей ---
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = header_font
        self.header_fill = header_fill

    def generate_xlsx_export(self, album_name, article_name=None, export_type='in_row', separator=', '):
        """Генерирует XLSX документ с ссылками на изображения
        Args:
            album_name (str): Название альбома
            article_name (str, optional): Название артикула. Если None - все артикулы альбома
            export_type (str): Тип экспорта - 'in_row' (в строку) или 'in_cell' (в ячейку)
            separator (str): Разделитель для варианта "в ячейку"
        Returns:
            tuple: (temp_file_path, filename) или (None, error_message)
        """
        try:
            logger.info(
                f"🔄 Начало генерации XLSX для альбома: {album_name}, артикул: {article_name}, тип: {export_type}")

            # Получаем данные из БД
            files_data = self._get_files_data(album_name, article_name)
            if not files_data:
                logger.warning(f"❌ Не найдены данные для экспорта: альбом={album_name}, артикул={article_name}")
                return None, "No data found for export"

            logger.info(f"📊 Найдено {len(files_data)} файлов для экспорта")

            # --- ИСПРАВЛЕНО: Workbook теперь доступен ---
            wb = Workbook()
            ws = wb.active
            ws.title = "Ссылки на изображения"

            # Применяем стили
            self._apply_header_styles(ws)

            # Генерируем содержимое в зависимости от типа экспорта
            if export_type == 'in_row':
                self._generate_in_row_export(ws, files_data)
                logger.debug("Сгенерирован экспорт типа 'в строку'")
            elif export_type == 'in_cell':
                self._generate_in_cell_export(ws, files_data, separator)
                logger.debug(f"Сгенерирован экспорт типа 'в ячейку' с разделителем: {repr(separator)}")
            else:
                logger.error(f"❌ Неизвестный тип экспорта: {export_type}")
                return None, f"Unknown export type: {export_type}"

            # Настраиваем ширину колонок
            self._auto_adjust_columns(ws)

            # Сохраняем во временный файл
            temp_file, temp_filename = self._save_to_temp_file(wb, album_name, article_name)
            filename = f"links_{album_name}{f'_{article_name}' if article_name else ''}.xlsx"

            # Логируем успешное создание файла
            logger.info(f"✅ Успешно создан XLSX файл: {filename}, временный путь: {temp_filename}")

            # Журналируем действие пользователя
            try:
                log_user_action(
                    action='export_xlsx',
                    resource_type='album' if not article_name else 'article',
                    resource_name=album_name if not article_name else f"{album_name}/{article_name}",
                    details={
                        'export_type': export_type,
                        'separator': separator,
                        'file_count': len(files_data),
                        'filename': filename,
                        'album_name': album_name,
                        'article_name': article_name
                    }
                )
            except Exception as log_error:
                logger.error(f"❌ Ошибка логирования действия экспорта: {log_error}")

            return temp_filename, filename

        except Exception as e:
            logger.error(f"❌ Ошибка генерации XLSX экспорта: {e}")
            # Логируем ошибку в действиях пользователя
            try:
                log_user_action(
                    action='export_xlsx_error',
                    resource_type='album' if not article_name else 'article',
                    resource_name=album_name if not article_name else f"{album_name}/{article_name}",
                    details={
                        'error': str(e),
                        'export_type': export_type,
                        'album_name': album_name,
                        'article_name': article_name
                    }
                )
            except Exception as log_error:
                logger.error(f"❌ Ошибка логирования ошибки экспорта: {log_error}")

            return None, f"Failed to create XLSX file: {str(e)}"

    def generate_csv_export(self, album_name, article_name=None):
        """
        Генерирует CSV документ с ссылками на изображения используя стандартный модуль csv
        Args:
            album_name (str): Название альбома
            article_name (str, optional): Название артикула
        Returns:
            tuple: (temp_file_path, filename) или (None, error_message)
        """
        try:
            logger.info(f"🔄 Начало генерации CSV для альбома: {album_name}, артикул: {article_name}")

            files_data = self._get_files_data(album_name, article_name)
            if not files_data:
                logger.warning(f"❌ Не найдены данные для CSV экспорта: альбом={album_name}, артикул={article_name}")
                return None, "No data found for export"

            logger.info(f"📊 Найдено {len(files_data)} файлов для CSV экспорта")

            # Группируем по артикулам и определяем максимальное количество ссылок
            articles_data = self._group_files_by_article(files_data)

            if not articles_data:
                logger.warning(f"❌ Не удалось сгруппировать данные для CSV экспорта")
                return None, "No data found for export"

            # Определяем максимальное количество ссылок
            max_links = max(len(links) for links in articles_data.values())

            # Создаем временный файл
            with tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.csv', encoding='utf-8',
                                             newline='') as tmp_file:
                # Используем стандартный модуль csv с правильными настройками
                writer = csv.writer(tmp_file, delimiter=',', quotechar='"',
                                    quoting=csv.QUOTE_MINIMAL, escapechar='\\')

                # Записываем заголовок с BOM для корректного отображения кириллицы в Excel
                tmp_file.write('\ufeff')

                # Создаем шапку как в XLSX "в строку" - артикул + нумерованные ссылки
                headers = ['Артикул'] + [f'Ссылка {i + 1}' for i in range(max_links)]
                writer.writerow(headers)

                # Записываем данные
                for article, links in articles_data.items():
                    # Создаем строку с артикулом и ссылками
                    row = [article] + links + [''] * (max_links - len(links))
                    writer.writerow(row)

            filename = f"links_{album_name}{f'_{article_name}' if article_name else ''}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

            logger.info(f"✅ Успешно создан CSV файл: {filename} с {len(articles_data)} записями")

            # Журналируем действие пользователя
            try:
                log_user_action(
                    action='export_csv',
                    resource_type='album' if not article_name else 'article',
                    resource_name=album_name if not article_name else f"{album_name}/{article_name}",
                    details={
                        'file_count': len(files_data),
                        'articles_count': len(articles_data),
                        'filename': filename,
                        'album_name': album_name,
                        'article_name': article_name
                    }
                )
            except Exception as log_error:
                logger.error(f"❌ Ошибка логирования действия CSV экспорта: {log_error}")

            return tmp_file.name, filename

        except Exception as e:
            logger.error(f"❌ Ошибка генерации CSV экспорта: {e}")
            # Логируем ошибку
            try:
                log_user_action(
                    action='export_csv_error',
                    resource_type='album' if not article_name else 'article',
                    resource_name=album_name if not article_name else f"{album_name}/{article_name}",
                    details={
                        'error': str(e),
                        'album_name': album_name,
                        'article_name': article_name
                    }
                )
            except Exception as log_error:
                logger.error(f"❌ Ошибка логирования ошибки CSV экспорта: {log_error}")

            return None, f"Failed to create CSV file: {str(e)}"

    # Остальные методы класса остаются без изменений...
    def _get_files_data(self, album_name, article_name):
        """Получает данные файлов из базы данных"""
        try:
            if article_name:
                results = db_manager.execute_query(
                    """SELECT filename, article_number, public_link 
                       FROM files 
                       WHERE album_name = %s AND article_number = %s 
                       ORDER BY article_number, filename""",
                    (album_name, article_name),
                    fetch=True
                )
            else:
                results = db_manager.execute_query(
                    """SELECT filename, article_number, public_link 
                       FROM files 
                       WHERE album_name = %s 
                       ORDER BY article_number, filename""",
                    (album_name,),
                    fetch=True
                )

            logger.debug(f"📋 Получено {len(results) if results else 0} записей из БД")
            return results if results else []

        except Exception as e:
            logger.error(f"❌ Ошибка получения данных из БД: {e}")
            return []

    def _group_files_by_article(self, files_data):
        """Группирует файлы по артикулам с правильной сортировкой"""

        def extract_suffix(filename):
            """Извлекает числовой суффикс из имени файла для сортировки"""
            import re
            match = re.search(r'(.+)_(\d+)(\.[^.]*)?$', filename)
            if match:
                return int(match.group(2))
            return 0

        # Сортируем результаты по артикулу и числовому суффиксу
        sorted_results = sorted(files_data, key=lambda x: (x['article_number'], extract_suffix(x['filename'])))

        # Группируем по артикулам
        articles_data = {}
        for row in sorted_results:
            article = row['article_number']
            if article not in articles_data:
                articles_data[article] = []
            articles_data[article].append(row['public_link'])

        logger.debug(f"📦 Сгруппировано по {len(articles_data)} артикулам")
        return articles_data

    def _apply_header_styles(self, worksheet):
        """Применяет стили для заголовков"""
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        self.header_font = header_font
        self.header_fill = header_fill

    def _generate_in_row_export(self, worksheet, files_data):
        """Генерирует экспорт типа 'в строку'"""
        # Группируем ссылки по артикулам
        articles_data = self._group_files_by_article(files_data)

        # Определяем максимальное количество ссылок
        max_links = max(len(links) for links in articles_data.values())

        # Создаем шапку
        headers = ['Артикул'] + [f'Ссылка {i + 1}' for i in range(max_links)]
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill

        # Заполняем данные
        for row, (article, links) in enumerate(articles_data.items(), 2):
            worksheet.cell(row=row, column=1, value=article)
            for col, link in enumerate(links, 2):
                worksheet.cell(row=row, column=col, value=link)

    def _generate_in_cell_export(self, worksheet, files_data, separator):
        """Генерирует экспорт типа 'в ячейку'"""
        # Группируем ссылки по артикулам
        articles_data = self._group_files_by_article(files_data)

        # Создаем шапку
        headers = ['Артикул', 'Ссылки']
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill

        # Заполняем данные
        for row, (article, links) in enumerate(articles_data.items(), 2):
            worksheet.cell(row=row, column=1, value=article)
            # Объединяем ссылки через разделитель
            links_text = separator.join(links)
            worksheet.cell(row=row, column=2, value=links_text)

    def _auto_adjust_columns(self, worksheet):
        """Автоматически настраивает ширину колонок"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def _save_to_temp_file(self, workbook, album_name, article_name):
        """Сохраняет workbook во временный файл"""
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            workbook.save(tmp_file.name)
            return tmp_file, tmp_file.name


def init_document_generator(base_url, upload_folder):
    """Инициализирует глобальный экземпляр генератора документов"""
    global document_generator
    document_generator = DocumentGenerator(base_url, upload_folder)
    logger.info("✅ Инициализирован генератор документов")
    return document_generator


def get_document_generator():
    """Возвращает глобальный экземпляр генератора документов"""
    global document_generator
    if document_generator is None:
        logger.error("❌ Генератор документов не инициализирован")
        raise RuntimeError("Document generator not initialized. Call init_document_generator first.")
    return document_generator
